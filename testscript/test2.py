import os
import cv2
import numpy as np
import pandas as data
import sys
import re
import pytesseract # this is tesseract module 
from pytesseract import Output
import matplotlib.pyplot as plt 
import glob
from PIL import Image
import argparse
import openpyxl
from pytesseract import Output
#parser = argparse.ArgumentParser(description='argument for image for pytesseract')
#parser.add_argument('-r','--row',type=str,required=True)
#parser.add_argument('-n','--number',type=str,required=True)
#parser.add_argument('-i', '--image', type=str, required=True,
 #                   help='Path for the image to be pass')


#args = parser.parse_args()

#IMAGE_NAME = args.i
def replace_str_index(text,index=0,replacement=''):
    return '%s%s%s'%(text[:index],replacement,text[index+1:])
def variance_of_laplacian(image):
	# compute the Laplacian of the image and then return the focus
	# measure, which is simply the variance of the Laplacian
	return cv2.Laplacian(image, cv2.CV_64F).var()
# construct the argument parse and parse the arguments
#IMAGE_NAME = '/home/cloudvms/ashwini/number.jpeg'
#imag = imageio.imread(IMAGE_NAME,as_gray=True)
#print(args.image)
image = cv2.imread('01610010512.0463266.jpg')
 
row = 0
col = 0
# By default worksheet names in the spreadsheet will be  
# Sheet1, Sheet2 etc., but we can also specify a name. 
wbkName = 'numbers.xlsx'
wbk = openpyxl.load_workbook(wbkName)
sheets = wbk.sheetnames
Sheet1 = wbk[sheets[0]]
gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
dst = cv2.resize(gray, None, fx = 2.5, fy = 2.5, interpolation = cv2.INTER_CUBIC)
fm = variance_of_laplacian(gray)
text = "Not Blurry"
# if the focus measure is less than the supplied threshold,
# then the image should be considered "blurry"
if fm < 3800.0:
	text = "Blurry"
	blur=cv2.GaussianBlur(dst,(1,1),0)
else:
        blur=cv2.GaussianBlur(dst,(7,7),0)
# show the image
print(text,fm)



ret,thresh=cv2.threshold(blur,0,255,cv2.THRESH_OTSU |cv2.THRESH_BINARY)
rect_kern=cv2.getStructuringElement(cv2.MORPH_RECT,(1,1))

dilation=cv2.dilate(thresh,rect_kern,iterations=1)
kernel = np.ones((5,5),np.uint8)
dilation = cv2.erode(dilation,kernel,iterations = 1)

contours,hierarchy=cv2.findContours(dilation,cv2.RETR_TREE,cv2.CHAIN_APPROX_SIMPLE)
sorted_contours=sorted(contours,key=lambda ctr:cv2.boundingRect(ctr)[0])
roi=0
plate_num=0
cout=811
coords = np.column_stack(np.where(thresh > 0))
angle = cv2.minAreaRect(coords)[-1]

if angle < -45:
	angle = -(90 + angle)
# otherwise, just take the inverse of the angle to make
# it positive
else:
	angle = -angle
(h, w) =thresh.shape[:2]
center = (w // 2, h // 2)
M = cv2.getRotationMatrix2D(center, angle, 1.0)
rotated = cv2.warpAffine(dilation, M, (w, h),
	flags=cv2.INTER_CUBIC, borderMode=cv2.BORDER_REPLICATE)
#cv2.putText(rotated, "Angle: {:.2f} degrees".format(angle),
#	(10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 255), 2)
print("[INFO] angle: {:.3f}".format(angle))
kernel = np.ones((1,1),np.uint8)
dilation1 = cv2.morphologyEx(dilation, cv2.MORPH_OPEN, kernel)    
dilation1 = cv2.morphologyEx(dilation, cv2.MORPH_CLOSE, kernel)
cropped=0
contours,h = cv2.findContours(dilation,1,2)
largest_rectangle = [0,0]
for cnt in contours:
    approx = cv2.approxPolyDP(cnt,0.01*cv2.arcLength(cnt,True),True)
    if len(approx)==4: #polygons with 4 points is what I need.
        area = cv2.contourArea(cnt)
        if area > largest_rectangle[0]:
            #find the polygon which has the largest size.
            largest_rectangle = [cv2.contourArea(cnt), cnt, approx]
x,y,w,h = cv2.boundingRect(largest_rectangle[1])
#crop the rectangle to get the number plate.
dilation=dilation[y:y+h,x:x+w]
cv2.imshow("ROI.jpg",dilation)
text = pytesseract.image_to_string(dilation,lang="eng", config='--psm 7  -c tessedit_char_whitelist=ABCDEFGHIJKLMOPQRSTUVWXYZ0123456789')
clean_text = re.sub('[\W_]+', '', text)
clean_text=replace_str_index(clean_text,0,'G')
print(clean_text)
cv2.waitKey(0)

#Sheet1 .cell(row=int(args.row), column=int(args.number)).value=args.image  
#Sheet1 .cell(row=int(args.row), column=int(args.number)+1).value=clean_text                 

#wbk.save(wbkName)
wbk.close
print("data entered")
#cv2.imshow("rotate",rotated)
#cv2.imshow("color",dilation)
#cv2.waitKey(0)
#cv.dis
