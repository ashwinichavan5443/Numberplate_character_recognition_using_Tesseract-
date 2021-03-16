import cv2 # For OpenCV modules (For Image I/O and Contour Finding)
import numpy as np # For general purpose array manipulation
import scipy.fftpack # For FFT2 
import pytesseract
import re
import openpyxl
from scipy.ndimage import interpolation as inter
#### imclearborder definition
def correct_skew(image, delta=1, limit=5):
    def determine_score(arr, angle):
        data = inter.rotate(arr, angle, reshape=False, order=0)
        histogram = np.sum(data, axis=1)
        score = np.sum((histogram[1:] - histogram[:-1]) ** 2)
        return histogram, score

    #gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    #thresh = cv2.threshold(image, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)[1] 

    scores = []
    angles = np.arange(-limit, limit + delta, delta)
    for angle in angles:
        histogram, score = determine_score(image, angle)
        scores.append(score)

    best_angle = angles[scores.index(max(scores))]

    (h, w) = image.shape[:2]
    center = (w // 2, h // 2)
    M = cv2.getRotationMatrix2D(center, best_angle, 1.0)
    rotated = cv2.warpAffine(image, M, (w, h), flags=cv2.INTER_CUBIC, \
              borderMode=cv2.BORDER_REPLICATE)

    return best_angle, rotated


def imclearborder(imgBW, radius):

    # Given a black and white image, first find all of its contours
    imgBWcopy = imgBW.copy()
    contours,hierarchy = cv2.findContours(imgBWcopy.copy(), cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)

    # Get dimensions of image
    imgRows = imgBW.shape[0]
    imgCols = imgBW.shape[1]    

    contourList = [] # ID list of contours that touch the border

    # For each contour...
    for idx in np.arange(len(contours)):
        # Get the i'th contour
        cnt = contours[idx]

        # Look at each point in the contour
        for pt in cnt:
            rowCnt = pt[0][1]
            colCnt = pt[0][0]

            # If this is within the radius of the border
            # this contour goes bye bye!
            check1 = (rowCnt >= 0 and rowCnt < radius) or (rowCnt >= imgRows-1-radius and rowCnt < imgRows)
            check2 = (colCnt >= 0 and colCnt < radius) or (colCnt >= imgCols-1-radius and colCnt < imgCols)

            if check1 or check2:
                contourList.append(idx)
                break

    for idx in contourList:
        cv2.drawContours(imgBWcopy, contours, idx, (0,0,0), -1)

    return imgBWcopy

#### bwareaopen definition
def bwareaopen(imgBW, areaPixels):
    # Given a black and white image, first find all of its contours
    imgBWcopy = imgBW.copy()
    contours,hierarchy = cv2.findContours(imgBWcopy.copy(), cv2.RETR_LIST, 
        cv2.CHAIN_APPROX_SIMPLE)

    # For each contour, determine its total occupying area
    for idx in np.arange(len(contours)):
        area = cv2.contourArea(contours[idx])
        if (area >= 0 and area <= areaPixels):
            cv2.drawContours(imgBWcopy, contours, idx, (0,0,0), -1)

    return imgBWcopy

#### Main program

# Read in image

sclst = ['AN', 'AP', 'AR', 'AS', 'BR', 'CH','CG','DD','DL','GJ','HR','HP','JK','JH','KA','KL','LA','LD','MP','MH','MN','ML','MZ','NL','OD','PY','PB','RJ','SK','TN','TS','TR','UP','UK','WB']
#string = 'properOCRImages/test08.jpg'
#string='numberplate/01610011516.8726232.jpg'
#string='proImg/test5.jpg'
string='test.jpg'
img1 = cv2.imread(string, 0)
img = cv2.imread(string, 0)
angle, img = correct_skew(img)
#cv2.imshow('test',img)
img = cv2.resize(img, None, fx = 2.5, fy = 2.5, interpolation = cv2.INTER_CUBIC)
#wbkName = 'testnumber.xlsx'
#wbk = openpyxl.load_workbook(wbkName)
#sheets = wbk.sheetnames
#Sheet1 = wbk[sheets[0]]
cv2.imshow("org",img)
# Number of rows and columns
rows = img.shape[0]
cols = img.shape[1]

# Remove some columns from the beginning and end

# Number of rows and columns
rows = img.shape[0]
cols = img.shape[1]

# Convert image to 0 to 1, then do log(1 + I)
imgLog = np.log1p(np.array(img, dtype="float") / 255)

# Create Gaussian mask of sigma = 10
M = 2*rows + 1
N = 2*cols + 1
sigma = 10
(X,Y) = np.meshgrid(np.linspace(0,N-1,N), np.linspace(0,M-1,M))
centerX = np.ceil(N/2)
centerY = np.ceil(M/2)
gaussianNumerator = (X - centerX)**2 + (Y - centerY)**2

# Low pass and high pass filters
Hlow = np.exp(-gaussianNumerator / (2*sigma*sigma))
Hhigh = 1 - Hlow

# Move origin of filters so that it's at the top left corner to
# match with the input image
HlowShift = scipy.fftpack.ifftshift(Hlow.copy())
HhighShift = scipy.fftpack.ifftshift(Hhigh.copy())

# Filter the image and crop
If = scipy.fftpack.fft2(imgLog.copy(), (M,N))
Ioutlow = scipy.real(scipy.fftpack.ifft2(If.copy() * HlowShift, (M,N)))
Iouthigh = scipy.real(scipy.fftpack.ifft2(If.copy() * HhighShift, (M,N)))

# Set scaling factors and add
gamma1 = 0.3
gamma2 = 1.5
Iout = gamma1*Ioutlow[0:rows,0:cols] + gamma2*Iouthigh[0:rows,0:cols]

# Anti-log then rescale to [0,1]
Ihmf = np.expm1(Iout)
Ihmf = (Ihmf - np.min(Ihmf)) / (np.max(Ihmf) - np.min(Ihmf))
Ihmf2 = np.array(255*Ihmf, dtype="uint8")

# Threshold the image - Anything below intensity 65 gets set to white

Ithresh = Ihmf2 < 65
Ithresh = 255*Ithresh.astype("uint8")
#cv2.imshow('test',Ithresh)
# Clear off the border.  Choose a border radius of 5 pixels
Iclear = imclearborder(Ithresh, 5)

# Eliminate regions that have areas below 120 pixels
Iopen = bwareaopen(Iclear, 120)

ret, thresh4 = cv2.threshold(Iopen,120,255,cv2.THRESH_OTSU )
thresh4=cv2.bitwise_not(thresh4)


print("[INFO] angle: {:.3f}".format(angle))
#cv2.imshow("rotate",rotated)
kernel = np.ones((5,5),np.uint8)
closing = cv2.morphologyEx(thresh4, cv2.MORPH_OPEN, kernel)
text = pytesseract.image_to_string(closing,lang="eng", config='--psm 3 --oem 2  -c tessedit_char_whitelist=ABCDEFGHIJKLMOPQRSTUVWXYZ0123456789')
clean_text = re.sub('[\W_]+', '', text)
sc=clean_text[0:2]
dc=clean_text[2:4]
mdl=clean_text[4:6]
ldl=clean_text[6:10]
#print(sc)
#print(dc)
#print(mdl)
#print(ldl)
final_strin= ""
only_alpha= ""
only_num = ""
only_mdl = ""
only_ldl = ""
if len(clean_text) == 10:
	print('yes')

	if sc.isalpha():
		if sc in sclst :
			#print("sc="+final_strin)	
			final_strin +=sc				
	
	else:
		for char in sc:
			if char.isdecimal():
						
				if char == '1':
					char="T"
				if char == '4':
					char="A"
				if char == '8':
					char="B"
				if char == '0':
					char="D"
				if char == '6':
					char = "G"
				if char == '5':
					char="S"
			only_num += char
		sc=only_num
		if sc in sclst :
			final_strin += sc
		else:
			print('No')

	if dc.isdecimal():
		final_strin +=dc
		#print("dc="+final_strin)			
	else:
		for char in dc:
			if char.isalpha():
						
				if char == 'S':
					char="5"
				if char == 'O':
					char="0"
				if char == 'Z':
					char="2"
				if char == 'A':
					char="4"
				if char == 'G':
					char = "6"
				if char == 'I':
					char = "1"
				if char == 'T':
					char = "1"
			only_mdl += char
		
		final_strin +=only_mdl
	if mdl.isalpha():
		final_strin +=mdl				
		
	else:
		for char in mdl:
			if char.isdecimal():
						
				if char == '1':
					char="T"
				if char == '4':
					char="A"
				if char == '8':
					char="B"
				if char == '0':
					char="Q"
				if char == '6':
					char = "G"
				if char == '5':
					char="S"
			only_mdl += char
		final_strin +=only_mdl
	if ldl.isdecimal():
		final_strin +=ldl
		print("ldl="+final_strin)			
	else:
		for char in ldl:
			if char.isalpha():
						
				if char == 'S':
					char="5"
				if char == 'O':
					char="0"
				if char == 'Q':
					char="0"
				if char == 'Z':
					char="2"
				if char == 'A':
					char="4"
				if char == 'G':
					char = "6"
				if char == 'E':
					char='6'
				if char == 'T':
					char='1'
				if char == 'B':
					char = '8'
				if char == 'F':
					char = '7'
				if char == 'I':
					char = '1'
				if char == 'C':
					char = '0'
				if char == 'Y':
					char = '9'
			only_ldl += char
		
		final_strin +=only_ldl	
else:
		final_strin="invalid number"
print("clean string="+final_strin)
#cv2.imwrite("cleanimg.jpg",thresh4)
print("ocr result="+clean_text)
# Show all images
r=28
c=1
#Sheet1 .cell(row=int(r), column=int(c)).value=string  
#Sheet1 .cell(row=int(r), column=int(c)+1).value=clean_text                 

#wbk.save(wbkName)
#wbk.close
cv2.imshow('Original Image', closing)
#cv2.imshow('Homomorphic Filtered Result', Ihmf2)
#cv2.imshow('Thresholded Result', Ithresh)
#cv2.imshow('Opened Result', Iopen)
cv2.waitKey(0)
cv2.destroyAllWindows()
